"""
export_import.py - Data export and import utilities
"""
import csv
import json
from datetime import datetime
from pathlib import Path
import sqlite3
from tkinter import filedialog, messagebox

try:
    import openpyxl
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ExportManager:
    """Handle data export to various formats"""
    
    def __init__(self, db_path="data/nano_erp.db"):
        self.db_path = Path(db_path)
    
    def export_customers_csv(self, output_path=None):
        """Export customers to CSV"""
        if output_path is None:
            output_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            if not output_path:
                return None
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM customers ORDER BY name')
        customers = cursor.fetchall()
        columns = [description[0] for description in cursor.description]
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            writer.writerows(customers)
        
        conn.close()
        return output_path
    
    def export_invoices_csv(self, output_path=None, start_date=None, end_date=None):
        """Export invoices to CSV"""
        if output_path is None:
            output_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            if not output_path:
                return None
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT i.*, c.name as customer_name
            FROM invoices i
            LEFT JOIN customers c ON i.customer_id = c.id
        '''
        
        params = []
        conditions = []
        
        if start_date:
            conditions.append("i.date >= ?")
            params.append(start_date)
        
        if end_date:
            conditions.append("i.date <= ?")
            params.append(end_date)
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        query += " ORDER BY i.date DESC"
        
        cursor.execute(query, params)
        invoices = cursor.fetchall()
        columns = [description[0] for description in cursor.description]
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            writer.writerows(invoices)
        
        conn.close()
        return output_path
    
    def export_products_excel(self, output_path=None):
        """Export products to Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showwarning("Excel Export", 
                                 "openpyxl is not installed. Install with: pip install openpyxl")
            return None
        
        if output_path is None:
            output_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not output_path:
                return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Products sheet
        cursor.execute('SELECT * FROM products ORDER BY name')
        products = cursor.fetchall()
        columns = [description[0] for description in cursor.description]
        
        # Write headers
        for col_num, column in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column)
        
        # Write data
        for row_num, row_data in enumerate(products, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_data)
        
        # Customers sheet
        ws2 = wb.create_sheet("Customers")
        cursor.execute('SELECT * FROM customers ORDER BY name')
        customers = cursor.fetchall()
        columns = [description[0] for description in cursor.description]
        
        for col_num, column in enumerate(columns, 1):
            ws2.cell(row=1, column=col_num, value=column)
        
        for row_num, row_data in enumerate(customers, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                ws2.cell(row=row_num, column=col_num, value=cell_data)
        
        conn.close()
        wb.save(output_path)
        return output_path
    
    def export_financial_report(self, output_path=None, month=None, year=None):
        """Export financial report"""
        if output_path is None:
            output_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            if not output_path:
                return None
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Get monthly summary
        query = '''
            SELECT 
                strftime('%Y-%m', date) as month,
                COUNT(*) as invoice_count,
                SUM(total) as total_revenue,
                SUM(CASE WHEN status = 'paid' THEN total ELSE 0 END) as paid_amount,
                SUM(CASE WHEN status = 'pending' THEN total ELSE 0 END) as pending_amount
            FROM invoices
            GROUP BY strftime('%Y-%m', date)
            ORDER BY month DESC
        '''
        
        cursor.execute(query)
        monthly_data = cursor.fetchall()
        
        # Get expense summary
        expense_query = '''
            SELECT 
                strftime('%Y-%m', date) as month,
                SUM(amount) as total_expenses
            FROM expenses
            GROUP BY strftime('%Y-%m', date)
            ORDER BY month DESC
        '''
        
        cursor.execute(expense_query)
        expense_data = cursor.fetchall()
        
        # Combine data
        report_data = []
        for month_row in monthly_data:
            month_str = month_row[0]
            expenses = next((e[1] for e in expense_data if e[0] == month_str), 0)
            profit = month_row[2] - expenses
            
            report_data.append({
                'Month': month_str,
                'Invoices': month_row[1],
                'Revenue': month_row[2],
                'Paid': month_row[3],
                'Pending': month_row[4],
                'Expenses': expenses,
                'Profit': profit
            })
        
        # Write to CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            if report_data:
                fieldnames = report_data[0].keys()
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(report_data)
        
        conn.close()
        return output_path
    
    def import_from_csv(self, file_path, table_name):
        """Import data from CSV to database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)
                
                # Create placeholders for SQL
                placeholders = ','.join(['?' for _ in headers])
                columns = ','.join(headers)
                
                # Check if table exists
                cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'")
                if not cursor.fetchone():
                    return False, f"Table '{table_name}' does not exist"
                
                # Import data
                for row in reader:
                    cursor.execute(f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})", row)
            
            conn.commit()
            return True, f"Successfully imported data to {table_name}"
            
        except Exception as e:
            conn.rollback()
            return False, f"Error importing data: {str(e)}"
        
        finally:
            conn.close()